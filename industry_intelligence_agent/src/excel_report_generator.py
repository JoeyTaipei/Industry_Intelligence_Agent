"""Excel workbook generator for the Streamlit MVP."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


logger = logging.getLogger(__name__)


def generate_excel_report(
    news_df: pd.DataFrame | None = None,
    annual_report_evidence_df: pd.DataFrame | None = None,
    kri_df: pd.DataFrame | None = None,
    dashboard_df: pd.DataFrame | None = None,
    chinese_summary: str | pd.DataFrame = "",
    output_path: str | Path = "data/reports/industry_intelligence_demo.xlsx",
    company_profile_df: pd.DataFrame | None = None,
) -> Path:
    """Create `industry_intelligence_demo.xlsx` with MVP output sheets."""
    _ = company_profile_df
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    summary_df = _summary_to_dataframe(chinese_summary)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        _ensure_dataframe(news_df).to_excel(writer, sheet_name="News_Articles", index=False)
        _ensure_dataframe(annual_report_evidence_df).to_excel(writer, sheet_name="Annual_Report_Evidence", index=False)
        _ensure_dataframe(kri_df).to_excel(writer, sheet_name="KRI_Evidence", index=False)
        _ensure_dataframe(dashboard_df).to_excel(writer, sheet_name="Dashboard_View", index=False, startrow=2)
        summary_df.to_excel(writer, sheet_name="Chinese_Summary", index=False)

    workbook = load_workbook(path)
    for sheet_name in workbook.sheetnames:
        header_row = 3 if sheet_name == "Dashboard_View" else 1
        if sheet_name == "Dashboard_View":
            _add_dashboard_note(workbook[sheet_name])
        _format_sheet(workbook[sheet_name], header_row=header_row)

    if "KRI_Evidence" in workbook.sheetnames:
        _add_conditional_formatting(workbook["KRI_Evidence"])
    if "Dashboard_View" in workbook.sheetnames:
        _add_conditional_formatting(workbook["Dashboard_View"], header_row=3)

    workbook.save(path)
    logger.info("Saved Excel report to %s", path)
    return path


def _add_dashboard_note(ws) -> None:
    note = "這份 Excel 展示如何用 Python 將年報與新聞資料轉成 KRI 風險證據，供顧問與 business user 進行人工覆核。"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws["A1"] = note
    ws["A1"].font = Font(bold=True, color="1F4E78")
    ws["A1"].fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")


def _format_sheet(ws, header_row: int = 1) -> None:
    ws.freeze_panes = f"A{header_row + 1}"
    if ws.max_row >= header_row and ws.max_column >= 1:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 55)


def _add_conditional_formatting(ws, header_row: int = 1) -> None:
    red_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    for column in range(1, ws.max_column + 1):
        header = str(ws.cell(row=header_row, column=column).value or "").lower()
        if "severity" not in header and "risk_level" not in header:
            continue
        letter = get_column_letter(column)
        cell_range = f"{letter}{header_row + 1}:{letter}{ws.max_row}"
        ws.conditional_formatting.add(cell_range, CellIsRule(operator="equal", formula=['"high"'], fill=red_fill))
        ws.conditional_formatting.add(cell_range, CellIsRule(operator="equal", formula=['"medium"'], fill=yellow_fill))
        ws.conditional_formatting.add(cell_range, CellIsRule(operator="equal", formula=['"low"'], fill=green_fill))
        ws.conditional_formatting.add(cell_range, CellIsRule(operator="equal", formula=['"High"'], fill=red_fill))
        ws.conditional_formatting.add(cell_range, CellIsRule(operator="equal", formula=['"Medium"'], fill=yellow_fill))
        ws.conditional_formatting.add(cell_range, CellIsRule(operator="equal", formula=['"Low"'], fill=green_fill))


def _summary_to_dataframe(summary: str | pd.DataFrame) -> pd.DataFrame:
    if isinstance(summary, pd.DataFrame):
        return summary
    lines = [line for line in str(summary or "").splitlines() if line.strip()]
    return pd.DataFrame({"summary_line": lines})


def _ensure_dataframe(value: Any) -> pd.DataFrame:
    if isinstance(value, pd.DataFrame):
        return value
    if isinstance(value, list):
        return pd.DataFrame(value)
    return pd.DataFrame()
